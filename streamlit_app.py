import streamlit as st
import pandas as pd
import os
from io import BytesIO
from openpyxl import load_workbook

# ✅ フォルダパス
TEMPLATE_FILE = "template.xlsx"
NG_LIST_FOLDER = "nglists"

# ✅ nglistsフォルダがなければ作成する
if not os.path.exists(NG_LIST_FOLDER):
    os.makedirs(NG_LIST_FOLDER)

# ✅ ページ設定
st.set_page_config(page_title="G-Change · Googleリスト整形＋NG除外", layout="centered")

st.title("🚗 G-Change：Googleリスト整形＋NG除外")
st.caption("ファイルをアップロード→クライアント選択→自動除外→ダウンロード")

# ✅ ファイルアップロード
uploaded_file = st.file_uploader("📄 整形対象のExcelファイルをアップロードしてください", type=["xlsx"])

# ✅ クライアント選択（nglists/を自動読み込み）
client_files = sorted([f for f in os.listdir(NG_LIST_FOLDER) if f.endswith(".xlsx")])
client_names = [os.path.splitext(f)[0] for f in client_files]

selected_client = None
ng_company_list = []
ng_tel_list = []

if client_names:
    selected_client = st.selectbox("📂 クライアント（NGリスト）を選択してください", client_names)

    if selected_client:
        nglist_path = os.path.join(NG_LIST_FOLDER, f"{selected_client}.xlsx")
        nglist_df = pd.read_excel(nglist_path)

        ng_company_list = nglist_df.iloc[:, 0].dropna().astype(str).str.strip().tolist()
        ng_tel_list = nglist_df.iloc[:, 1].dropna().astype(str).str.replace("ー", "-").str.replace("−", "-").str.replace("―", "-").str.strip().tolist()

        with st.expander("🔍 NGリストの中身を確認する"):
            st.dataframe(nglist_df)

# ✅ メイン処理
if uploaded_file:
    file_name = uploaded_file.name
    df_raw = pd.read_excel(uploaded_file)

    # 🔵 並べ替え処理（もし必要なら）
    if list(df_raw.columns) != ['企業名', '業種', '住所', '電話番号']:
        st.warning("⚠️ 縦型リストを想定して整形を進めます！")
        df_raw.columns = ['企業名']
        df_raw['業種'] = df_raw['企業名'].shift(-2)
        df_raw['住所'] = df_raw['企業名'].shift(-1)
        df_raw['電話番号'] = df_raw['企業名']
        df_raw = df_raw.iloc[3::4].reset_index(drop=True)
        df_raw = df_raw[['企業名', '業種', '住所', '電話番号']]

    # 🔵 電話番号フォーマット補正
    df_raw['電話番号'] = df_raw['電話番号'].astype(str).str.replace("ー", "-").str.replace("−", "-").str.replace("―", "-").str.strip()

    # 🔵 NG除外処理
    if selected_client:
        df_filtered = df_raw[
            ~df_raw['企業名'].astype(str).str.strip().isin(ng_company_list) &
            ~df_raw['電話番号'].astype(str).str.strip().isin(ng_tel_list)
        ]
    else:
        df_filtered = df_raw

    # 📥 整形結果をダウンロード
    if st.button("▶️ 整形＋NG除外を実行する"):
        output = BytesIO()

        wb = load_workbook(TEMPLATE_FILE)
        ws = wb['入力マスター']

        for idx, row in df_filtered.iterrows():
            ws.cell(row=idx+2, column=2, value=row['企業名'])
            ws.cell(row=idx+2, column=3, value=row['業種'])
            ws.cell(row=idx+2, column=4, value=row['住所'])
            ws.cell(row=idx+2, column=5, value=row['電話番号'])

        wb.save(output)
        output.seek(0)

        st.success(f"✅ 完了しました！企業数：{len(df_filtered)}件")
        st.download_button(
            label="📥 ダウンロードする",
            data=output,
            file_name=f"{os.path.splitext(file_name)[0]}：リスト（NG除外済）.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )